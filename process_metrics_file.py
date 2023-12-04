"""
I want to create a function that converts the given MetricsOutput.tsv
file into a MetricsOutput.xlsx file that does the following changes

Changes:
    - from rows 16-16, move data 2 cells to the right from column B
    - from row 17, higlight in red any cell that says FALSE
    - set a variable for threshold cells in B&C 
        - [DNA Library QC Metrics] 23-24
        - [DNA Library QC Metrics for Small Variant Calling and TMB] 28-30
        - [DNA Library QC Metrics for MSI] 34
        - [DNA Library QC Metrics for CNV] 37-38
        - [RNA Library QC Metrics] 60-62
        - [RNA Expanded Metrics] 66-69
    - highlight the subsequent cells if outside thresholds
"""
import argparse
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, Rule


def parse_args() -> argparse.Namespace:
    """
    Parse command line arguments

    Returns:
        args : Namespace
          Namespace of passed command line argument inputs
    """
    parser = argparse.ArgumentParser(description='MetricsOutput.tsv converter and editor')

    parser.add_argument('tsv_input', type=str,
                        help='filepath to MetricsOutput.tsv file')
    parser.add_argument('-o', '--output_filename', type=str, default='MetricsOutput.xlsx',
                        help='OPTIONAL: Output filename, default set to MetricsOutput.xlsx')
    args = parser.parse_args()

    return args


def tsv_to_excel(input_filepath, output_filepath):
    """
    Function which converts given .tsv file into excel file

    Args:
        input_filepath (str)_
        output_filepath (str)

    """
    workbook_object = openpyxl.Workbook()
    worksheet = workbook_object.active

    with open(input_filepath, 'r', encoding='UTF-8') as csvfile:
        read_tsv = csv.reader(csvfile, delimiter = '\t')
        for row in read_tsv:
            converted_row = []
            for item in row:
                try:
                    # Convert to float if possible
                    converted_row.append(float(item))
                except ValueError:
                    try:
                        # Convert to int if float not possible
                        converted_row.append(int(item))
                    except ValueError:
                        converted_row.append(item)
            worksheet.append(converted_row)

    workbook_object.save(output_filepath)


def edit_excel(excel_file):
    """
    Edits the given excel file as wanted

    Args:
        excel_file (str): filepath of the excel file.
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Sheet']
    # Change name of sheet from excel workbook
    ws.title = "MetricsOutput"

    # Openpyxl.utils fuction get_column_letter() prints the letter of an column number
    # e.g: get_column_letter(23) == 'W'
    max_column = get_column_letter(ws.max_column)

    # From rows 16-19, move data 2 cells to the right from column B
    ws.move_range(f'B16:{max_column}19', cols=2)

    # Colour cells in red if cells from row 17 contain FALSE
    red_text = Font(name='Calibri', color="9C0006")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="FALSE", dxf=dxf)
    ws.conditional_formatting.add(f'D17:{max_column}17', rule)

    # Colour cells in red if contamination score and contamination p value are outside USL
    CONTAMINATION_SCORE_ROW = 23
    usl_score = ws[f'C{CONTAMINATION_SCORE_ROW}'].value

    CONTAMINATION_P_VALUE_ROW = 24
    usl_p_value = ws[f'C{CONTAMINATION_P_VALUE_ROW}'].value

    for col in list(range(4,(ws.max_column))):
        letter = get_column_letter(col)
        score_to_compare = ws[f'{letter}{CONTAMINATION_SCORE_ROW}'].value
        try:
            if usl_score < score_to_compare:
                p_value_to_compare = ws[f'{letter}{CONTAMINATION_P_VALUE_ROW}'].value
                if usl_p_value < p_value_to_compare:
                    ws[f'{letter}{CONTAMINATION_SCORE_ROW}'].fill = red_fill
                    ws[f'{letter}{CONTAMINATION_P_VALUE_ROW}'].fill = red_fill
                    ws[f'{letter}{CONTAMINATION_SCORE_ROW}'].font = red_text
                    ws[f'{letter}{CONTAMINATION_P_VALUE_ROW}'].font = red_text
        except TypeError:
            pass

    # Colour cells in red for the rows of interest
    rows_to_format = [28, 29, 30, 34, 37, 38, 60,
                      61, 62, 66, 67, 68, 69]
    for row in rows_to_format:
        LSL = ws[f'B{row}'].value
        USL = ws[f'C{row}'].value
        if LSL == 'NA' and USL == 'NA':
            pass
        elif LSL == 'NA':
            operator = 'greaterThan'
            formula = [USL]
        elif USL == 'NA':
            operator = 'lessThan'
            formula = [LSL]
        else:
            operator = 'notBetween'
            formula = [LSL, USL]
        rule = CellIsRule(operator=operator, formula=formula,
                          stopIfTrue=False, fill=red_fill, font=red_text)
        ws.conditional_formatting.add(f'D{row}:{max_column}{row}', rule)

    wb.save(excel_file)


def main():
    """
    Main entry points to run the script. Creates an .xlsx file when run correctly
    """
    args = parse_args()

    tsv_to_excel(args.tsv_input, args.output_filename)
    edit_excel(args.output_filename)

    print(args.tsv_input)
    print(args.output_filename)

if __name__ == '__main__':
    main()
