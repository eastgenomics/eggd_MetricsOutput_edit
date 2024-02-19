"""
Python app that takes the  MetricsOutput.tsv file and creates the
MetricsOutput.xlsx file that does the following changes

Changes:
    - from rows 16-19, move data 2 cells to the right from column B
    - from row 17, higlight in red any cell that says FALSE
    - set a variable for threshold cells in columns B and C
        - [DNA Library QC Metrics] 23-24
        - [DNA Library QC Metrics for Small Variant Calling and TMB] 28-30
        - [DNA Library QC Metrics for MSI] 34
        - [DNA Library QC Metrics for CNV] 38-39
        - [RNA Library QC Metrics] 60-62
        - [RNA Expanded Metrics] 66-69
    - highlight the subsequent cells if outside thresholds

"""

import argparse
import csv
from os import remove
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule

RED_TEXT = Font(name='Calibri', color="9C0006")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")


class Excel():
    """
    Functions for formatting excel data with the appropriate style and
    writing output file

    Attributes
    ----------
    """
    def __init__(self, file) -> None:
        print(f"Editing excel file {file}")
        # String with excel filename
        self.file = file
        # Workbook of excel file
        self.wb = openpyxl.load_workbook(self.file)
        # Worksheet of interest from excel file
        self.ws = self.wb['Sheet']
        # Change the worksheet name to MetricsOutput
        self.ws.title = "MetricsOutput"
        # Set a pandas dataframe which reflects the contents of the worksheet
        self.df = pd.DataFrame(self.ws.values)

    def modify(self) -> None:
        """
        Calls all methods in excel() to modify and generate output file
        with desired changes
        """
        self.move_rows()
        self.mark_false()
        self.mark_contamination_metrics()
        self.mark_other_metrics()
        self.wb.save(self.file)
        print(f"Done! {self.file} file generated")

    def move_rows(self) -> None:
        """
        Excel method which modifies its input by moving the set of rows
        from 16 to 19 by 2 columns towards the right.

        This is to shift the [Analysis status] table to allign sample metrics
        with the rest of the tables from the excel file.
        """
        # Obtain the last column letter of the excel file and store it
        # in a variable
        max_column = get_column_letter(self.ws.max_column - 2)

        self.ws.move_range(f'B16:{max_column}19', cols=2)
        self.df = pd.DataFrame(self.ws.values)

    def mark_false(self) -> None:
        """
        Mark in red all cells with string FALSE.
        An error will be raised if the string is found outside of row 17

        """
        string_to_find = "FALSE"
        # store a list of cell indices where the contain a string equal to
        # "FALSE"
        false_cells_indices = self.df.stack().index[self.df.stack() ==
                                                    string_to_find]
        # Mark every cell from the list in red
        for idx in false_cells_indices:
            row = idx[0]+1
            if row != 17:
                remove(self.file)
                raise TypeError("FALSE string found outside expected row.")
            excel_column = get_column_letter(idx[1]+1)
            self.ws[f'{excel_column}{row}'].fill = RED_FILL
            self.ws[f'{excel_column}{row}'].font = RED_TEXT

    def mark_contamination_metrics(self) -> None:
        """
        Mark in red the DNA Library QC metrics when values
        exceed the guidelines on all metrics from the elements_to_find list
        """
        elements_to_find = ["CONTAMINATION_SCORE (NA)",
                            "CONTAMINATION_P_VALUE (NA)"]

        # Set a loop for each sample
        for sample_col_index in range(3, len(self.df.columns)):
            # Set loop for each metric
            for element in elements_to_find:
                indices = self.df.stack().index[self.df.stack() == element]
                for idx in indices:
                    row = idx[0]
                    first_column = idx[1]
                    LSL_column_index = first_column + 1
                    USL_column_index = first_column + 2

                    value_to_compare = self.df.loc[row][sample_col_index]

                    # Ensure that each string from the elements_to_find are
                    # found in the first column
                    if first_column == 0:
                        LSL = self.df.loc[row][LSL_column_index]
                        USL = self.df.loc[row][USL_column_index]

                # Store a boolean for each sample and set to false when any
                # metric does not exceed the USL and LSL thresholds
                if value_to_compare < LSL or value_to_compare > USL:
                    sample_to_highlight = True
                else:
                    sample_to_highlight = False
                    break

            # If boolean remains true, highlight every value for sample
            if sample_to_highlight is True:
                for element in elements_to_find:
                    indices = self.df.stack().index[self.df.stack() == element]
                    for idx in indices:
                        excel_row = idx[0]+1
                        excel_column = get_column_letter(sample_col_index+1)
                        self.ws[f'{excel_column}{excel_row}'].fill = RED_FILL
                        self.ws[f'{excel_column}{excel_row}'].font = RED_TEXT

    def mark_other_metrics(self):
        """
        Mark in red the DNA Library QC metrics when
        values exceed the guidelines
        """

        max_column = get_column_letter(self.ws.max_column)
        metrics_to_find = ['MEDIAN_INSERT_SIZE (bp)',
                           'MEDIAN_EXON_COVERAGE (Count)',
                           'PCT_EXON_50X (%)',
                           'USABLE_MSI_SITES (Count)',
                           'COVERAGE_MAD (Count)',
                           'MEDIAN_BIN_COUNT_CNV_TARGET (Count)',
                           'MEDIAN_CV_GENE_500X (NA)',
                           'TOTAL_ON_TARGET_READS (NA)',
                           'MEDIAN_INSERT_SIZE (NA)',
                           'PCT_CHIMERIC_READS (NA)',
                           'PCT_ON_TARGET_READS (NA)',
                           'SCALED_MEDIAN_GENE_COVERAGE (NA)',
                           'TOTAL_PF_READS (NA)']

        # Search for the cell location for every metric from metrics list
        for metric in metrics_to_find:
            indices = self.df.stack().index[self.df.stack() == metric]

            # Once cell found, assign variables to USL and LSL guidelines
            for idx in indices:
                row = idx[0]
                LSL = self.df.loc[row][idx[1]+1]
                USL = self.df.loc[row][idx[1]+2]

            # Select the appropriate operator and formula based on
            # the values of LSL and USL
            if LSL == 'NA' and USL == 'NA':
                pass
            elif LSL == 'NA':
                operator = 'greaterThan'
                formula = [USL]
            elif USL == 'NA':
                operator = 'lessThan'
                formula = [LSL]
            else:
                operator = 'notBetween'
                formula = [LSL, USL]

            # for every row, apply selected operator and formula
            # to highlight cells
            rule = CellIsRule(operator=operator, formula=formula,
                              stopIfTrue=False, fill=RED_FILL,
                              font=RED_TEXT)
            self.ws.conditional_formatting.add(f'D{row+1}:{max_column}{row+1}',
                                               rule)


def parse_args() -> argparse.Namespace:
    """
    Parse command line arguments

    Returns:
        args : Namespace
          Namespace of passed command line argument inputs
    """
    parser = argparse.ArgumentParser(description='MetricsOutput.tsv \
                                     converter and editor')

    parser.add_argument('tsv_input', type=str,
                        help='filepath to MetricsOutput.tsv file')
    parser.add_argument('-o', '--output_filename', type=str,
                        default='MetricsOutput.xlsx',
                        help='OPTIONAL: Output filename, default set to \
                              MetricsOutput.xlsx')
    args = parser.parse_args()

    return args


def tsv_to_excel(input_filepath, output_filepath):
    """
    Function which creates an excel file from a .tsv file.

    Args:
        input_filepath (str)
        output_filepath (str)

    """
    workbook_object = openpyxl.Workbook()
    worksheet = workbook_object.active

    # Convert .tsv file to excel
    with open(input_filepath, 'r', encoding='UTF-8') as csvfile:
        read_tsv = csv.reader(csvfile, delimiter='\t')
        for row in read_tsv:
            converted_row = []
            for item in row:
                try:
                    # Convert to float if possible
                    converted_row.append(float(item))
                except ValueError:
                    try:
                        # Convert to int if float not possible
                        converted_row.append(int(item))
                    except ValueError:
                        converted_row.append(item)
            worksheet.append(converted_row)

    workbook_object.save(output_filepath)


def main():
    """
    Main entry points to run the script. Creates an .xlsx file when run
    correctly
    """
    args = parse_args()

    tsv_to_excel(args.tsv_input, args.output_filename)

    excel_file = Excel(args.output_filename)
    excel_file.modify()

    print(args.tsv_input)
    print(args.output_filename)


if __name__ == '__main__':
    main()
